#!/usr/bin/env python3
# -*- coding: utf-8 -*-

""" get creditcode """

__author__ = '陈晓炜'

import openpyxl
from urllib import request
import urllib.parse
import json
import time


def rwExcel(readpath, writepath):
    wb = openpyxl.load_workbook(readpath)
    sheet = wb.worksheets[1]  # 证照信息页
    start_row = 5  # 起始行
    row_len = sheet.max_row  # 行数
    for row in range(start_row, row_len + 1):
        cell = sheet.cell(row=row, column=6).value  # 读取企业名称 查列数
        sheet.cell(row=row, column=8).value = search(cell)  # 写入统一社会信用代码 写列数
        time.sleep(1.5)  # 暂停时间 单位秒
        wb.save(writepath)  # 保存


def search(keyword):
    url = "https://www.creditchina.gov.cn/api/credit_info_search?keyword=" \
          + urllib.parse.quote(keyword)
    page = ''
    while page == '':
        xml_info = request.urlopen(url)
        page = xml_info.read().decode('utf-8')
    json_dic = json.loads(page)  # JSON转换为字典
    try:
        data = json_dic['data']
    except KeyError:
        print("访问受限"+" "+keyword)
        exit()
    else:
        totalCount = data['totalCount']
        if totalCount != 0:
            encryStr = data['results'][0]['encryStr']  # 获取验证码
        else:
            encryStr = 0
        if encryStr != 0:
            url_sec = "https://www.creditchina.gov.cn/api/credit_info_detail?encryStr=" \
                      + urllib.parse.quote(encryStr)
            xml_info = request.urlopen(url_sec)
            page_sec = xml_info.read().decode('utf-8')
            try:
                json_dic_sec = json.loads(page_sec)  # JSON转换为字典
                result = json_dic_sec['result']
            except KeyError:
                print("json格式错误"+" "+keyword)
                creditCode = ''
            else:
                try:
                    result['creditCode']
                except TypeError:
                    print("查无统一社会信用代码" + " " + keyword)
                    creditCode = ''
                else:
                    creditCode = result['creditCode']  # 获取统一社会信用代码
        else:
            print("查无统一社会信用代码"+" "+keyword)
            creditCode = ''
    return creditCode


if __name__ == '__main__':
    read_path = 'F:\新建文件夹\市食品药品监督管理局-中华人民共和国药品经营质量管理规范认证证书-V1.0.xlsx'  # 读取的文件路径
    write_path = 'F:\新建文件夹\测试2.xlsx'  # 写入的文件路径
    rwExcel(read_path, write_path)
